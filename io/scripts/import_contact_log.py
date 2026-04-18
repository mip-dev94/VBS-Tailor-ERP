# -*- coding: utf-8 -*-
"""Import sheet "Liên hệ khách hàng" từ Nhật ký Excel → vbs.contact.log.

Usage:
    docker exec -i custom_addons-odoo-1 odoo shell \
        -c /etc/odoo/odoo.conf -d VBS_ERP --no-http \
        < io/scripts/import_contact_log.py

Idempotency: lookup theo (date_contact, partner_id, noi_dung). Trùng → skip.
"""
import openpyxl
exec(open('/mnt/io/scripts/_lib.py', encoding='utf-8').read())

INPUT_FILE = '/mnt/io/input/nhat_ky.xlsx'
SHEET_NAME = 'Liên hệ khách hàng'
HEADER_ROW = 2       # row 2 là header (row 1 trống)
DATA_START_ROW = 3
DRY_RUN = False

wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws = wb[SHEET_NAME]

header = [str(c.value).strip() if c.value else '' for c in ws[HEADER_ROW]]
print('HEADER:', header[:10])


def col(name):
    for i, h in enumerate(header):
        if h.lower() == name.lower():
            return i
    return -1


IDX = {
    'date': col('Ngày liên hệ'),
    'customer': col('Tên khách'),
    'content': col('Nội dung'),
    'result': col('Kết quả'),
    'status': col('Tình trạng liên hệ'),
}
print('Column map:', IDX)

created = skipped = failed = 0
Log = env['vbs.contact.log']

for row_idx, row in enumerate(
    ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
):
    try:
        customer = row[IDX['customer']] if IDX['customer'] >= 0 else None
        date = to_date(row[IDX['date']]) if IDX['date'] >= 0 else False
        if not customer or not str(customer).strip() or not date:
            continue

        partner = find_or_create_partner(env, customer)
        noi_dung = (row[IDX['content']] if IDX['content'] >= 0 else '') or ''
        ket_qua = (row[IDX['result']] if IDX['result'] >= 0 else '') or ''
        status = map_value(
            row[IDX['status']] if IDX['status'] >= 0 else None,
            CONTACT_STATUS_MAP, default='cho_lien_he', label='tinh_trang',
        )

        # Idempotency: skip nếu đã có (same day + partner + content)
        dup = Log.search([
            ('date_contact', '=', date),
            ('partner_id', '=', partner.id),
            ('noi_dung', '=', noi_dung),
        ], limit=1)
        if dup:
            skipped += 1
            continue

        Log.create({
            'date_contact': date,
            'partner_id': partner.id,
            'noi_dung': str(noi_dung)[:255],
            'ket_qua': str(ket_qua)[:255],
            'tinh_trang': status,
        })
        created += 1
    except Exception as e:
        failed += 1
        print(f'  [FAIL] row {row_idx}: {e}')

wb.close()
print(f'\n=== KẾT QUẢ ===')
print(f'  created: {created}')
print(f'  skipped (dup): {skipped}')
print(f'  failed:  {failed}')

if DRY_RUN:
    env.cr.rollback()
    print('\n>>> DRY_RUN=True — rollback')
else:
    env.cr.commit()
    print('\n>>> committed')
