# -*- coding: utf-8 -*-
"""Import 1 file DV01A (vd: 20251130-HaNoi-DV01A.xlsx) → 1 vbs.fabric.order
chứa N lines.

Usage:
    docker exec -i custom_addons-odoo-1 odoo shell \
        -c /etc/odoo/odoo.conf -d VBS_ERP --no-http \
        < io/scripts/import_dv01a.py

Đổi INPUT_FILE nếu file khác tên. Script tự phát hiện header dựa vào cột 'STT'.
"""
import openpyxl
exec(open('/mnt/io/scripts/_lib.py', encoding='utf-8').read())

INPUT_FILE = '/mnt/io/input/dv01a.xlsx'
SHEET_NAME = 'Mẫu 1'
DRY_RUN = False

# Mapping cho các Selection trên vbs.fabric.order.line (tự đóng khung trong script)
GARMENT_DESC_MAP = {
    '01 sơ mi': 'so_mi',
    '01 áo khoác lẻ': 'ao_khoac_le',
    '01 quần lẻ': 'quan_le',
    '01 bộ comple 2 mảnh': 'bo_comple_2',
    '01 bộ comple 3 mảnh': 'bo_comple_3',
    '01 gile': 'gile',
    '01 áo khoác tạo kiểu': 'ao_khoac_tao_kieu',
    '01 lót áo khoác': 'lot_ao_khoac',
    '01 bộ cúc': 'bo_cuc',
}
PATTERN_MAP_LINE = {
    'trơn': 'tron',
    'kẻ ô vừa': 'ke_o_vua',
    'kẻ ô to (trên 10cm)': 'ke_o_to',
    'kẻ ô to': 'ke_o_to',
    'kẻ dọc': 'ke_doc',
    'kẻ dọc to': 'ke_doc_to',
}
BUTTON_ROW_MAP = {'một hàng cúc': 'mot_hang', 'hai hàng cúc': 'hai_hang'}
POCKET_MAP = {'ốp': 'op', 'nắp': 'nap'}
LINING_MAP = {'lót cả': 'lot_ca', 'lót nửa': 'lot_nua', 'không lót': 'khong_lot'}
CUFF_MAP = {'có lv': 'co_lv', 'không lv': 'khong_lv'}


wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws = wb[SHEET_NAME]

# Tìm header row: row có ô đầu tiên = 'STT'
header_row = None
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row and row[0] and str(row[0]).strip().lower() == 'stt':
        header_row = i
        break
if not header_row:
    raise RuntimeError('Không tìm được header row "STT" trong sheet')
data_start = header_row + 2  # +1 header, +1 sub-header row
print(f'Header row={header_row}, data bắt đầu từ row={data_start}')

# Đọc header (gộp 2 row header + sub-header)
h1 = [str(c.value).strip() if c.value else '' for c in ws[header_row]]
h2 = [str(c.value).strip() if c.value else '' for c in ws[header_row + 1]]
header = [h2[i] if h2[i] else h1[i] for i in range(max(len(h1), len(h2)))]
print('Header (merged):', header[:20])


def col(name):
    for i, h in enumerate(header):
        if h and h.lower() == name.lower():
            return i
    return -1


IDX = {
    'branch': col('Chi nhánh'),   # có thể ignore sau khi đã đi single-branch
    'ma_do': col('Mã đồ'),
    'sapo': col('Mã sapo'),
    'customer': col('Tên khách hàng'),
    'brand': col('Hãng vải'),
    'code': col('Mã vải'),
    'desc': col('Mô tả đồ đặt'),
    'qty': col('Khối lượng'),
    'unit': col('Đơn vị'),
    'pattern': col('Vải kẻ'),
    'dai_ao': col('Dài áo'),
    'dai_tay': col('Dài tay áo'),
    'button_row': col('Số hàng cúc của áo'),
    'pocket': col('Túi áo'),
    'lining': col('Lót áo'),
    'dai_quan': col('Dài quần'),
    'cuff': col('Lơ Vê'),
    'date': col('Ngày lên đơn'),
}
print('Column map:', IDX)

# Lấy ngày lên đơn từ row đầu data (tất cả dòng thường cùng ngày)
first_data_row = list(ws.iter_rows(min_row=data_start, max_row=data_start, values_only=True))[0]
date_order = to_date(first_data_row[IDX['date']]) if IDX['date'] >= 0 else False

Order = env['vbs.fabric.order']
order = Order.create({
    'date_order': date_order or False,
    'state': 'draft',
})
print(f'Tạo phiếu DV01: {order.name} (id={order.id})')

Line = env['vbs.fabric.order.line']
lines_created = failed = 0
for row_idx, row in enumerate(
    ws.iter_rows(min_row=data_start, values_only=True), start=data_start
):
    try:
        stt = row[0]
        if stt is None or str(stt).strip() == '':
            continue  # row trống = kết thúc
        customer = row[IDX['customer']] if IDX['customer'] >= 0 else None
        if not customer:
            continue
        partner = find_or_create_partner(env, customer)

        desc = map_value(
            row[IDX['desc']] if IDX['desc'] >= 0 else None,
            GARMENT_DESC_MAP, default=False, label='garment_desc',
        )
        pattern = map_value(
            row[IDX['pattern']] if IDX['pattern'] >= 0 else None,
            PATTERN_MAP_LINE, default=False, label='pattern',
        )

        vals = {
            'order_id': order.id,
            'partner_id': partner.id,
            'sapo_code': str(row[IDX['sapo']]).strip() if (IDX['sapo'] >= 0 and row[IDX['sapo']]) else '',
            'garment_ref': str(row[IDX['ma_do']]).strip() if (IDX['ma_do'] >= 0 and row[IDX['ma_do']]) else '',
            'fabric_brand': str(row[IDX['brand']]).strip()[:64] if (IDX['brand'] >= 0 and row[IDX['brand']]) else '',
            'fabric_code': str(row[IDX['code']]).strip()[:64] if (IDX['code'] >= 0 and row[IDX['code']]) else '',
            'garment_desc': desc,
            'pattern': pattern,
            'quantity': float(row[IDX['qty']]) if (IDX['qty'] >= 0 and row[IDX['qty']]) else 0.0,
            'dai_ao': str(row[IDX['dai_ao']]).strip()[:32] if (IDX['dai_ao'] >= 0 and row[IDX['dai_ao']]) else '',
            'dai_tay_ao': str(row[IDX['dai_tay']]).strip()[:32] if (IDX['dai_tay'] >= 0 and row[IDX['dai_tay']]) else '',
            'dai_quan': str(row[IDX['dai_quan']]).strip()[:32] if (IDX['dai_quan'] >= 0 and row[IDX['dai_quan']]) else '',
            'button_row': map_value(row[IDX['button_row']] if IDX['button_row'] >= 0 else None, BUTTON_ROW_MAP, default=False),
            'pocket': map_value(row[IDX['pocket']] if IDX['pocket'] >= 0 else None, POCKET_MAP, default=False),
            'lining': map_value(row[IDX['lining']] if IDX['lining'] >= 0 else None, LINING_MAP, default=False),
            'cuff': map_value(row[IDX['cuff']] if IDX['cuff'] >= 0 else None, CUFF_MAP, default=False),
        }
        Line.create(vals)
        lines_created += 1
    except Exception as e:
        failed += 1
        print(f'  [FAIL] row {row_idx}: {e}')

wb.close()
print(f'\n=== KẾT QUẢ ===')
print(f'  order: {order.name}')
print(f'  lines created: {lines_created}')
print(f'  failed:        {failed}')

if DRY_RUN:
    env.cr.rollback()
    print('\n>>> DRY_RUN=True — rollback')
else:
    env.cr.commit()
    print('\n>>> committed')
