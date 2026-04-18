# -*- coding: utf-8 -*-
"""Import sheet "Xuất nhập đồ(new)" từ Nhật ký Excel → vbs.garment.

Usage:
    docker exec -i custom_addons-odoo-1 odoo shell \
        -c /etc/odoo/odoo.conf -d VBS_ERP --no-http \
        < io/scripts/import_garment_log.py

Cấu hình: chỉnh INPUT_FILE + SHEET_NAME ở dưới nếu cần.
Idempotency: lookup theo (partner_name, detail) — nếu khớp → update, không khớp → create.
             Set DRY_RUN = True để preview trước khi commit.
"""
import openpyxl
from datetime import datetime
exec(open('/mnt/io/scripts/_lib.py', encoding='utf-8').read())

INPUT_FILE = '/mnt/io/input/nhat_ky.xlsx'
SHEET_NAME = 'Xuất nhập đồ(new)'
HEADER_ROW = 1       # 1-indexed: row 1 là header
DATA_START_ROW = 3   # 1-indexed: row 3 là data đầu tiên (row 2 là "*Diễn giải")
DRY_RUN = False      # True = không commit, chỉ preview

wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws = wb[SHEET_NAME]

# Đọc header để map column index theo tên
header = [str(c.value).strip() if c.value else '' for c in ws[HEADER_ROW]]
print('HEADER:', header[:25])


def col(name):
    """Tìm column index (0-based) của tên cột; -1 nếu không có."""
    for i, h in enumerate(header):
        if h.lower() == name.lower():
            return i
    return -1


IDX = {
    'customer': col('Tên khách'),
    'type': col('Loại đồ'),
    'detail': col('Chi tiết đồ'),
    'state': col('Tình trạng'),
    'location': col('Vị trí'),
    'date_return': col('Ngày trả khách'),
}
# 8 round LCH/LX
LCH_IDX = [col(f'LCH{i}') for i in range(1, 9)]
LX_IDX = [col(f'LX{i}') for i in range(1, 9)]

print('Column map:', IDX)
print('LCH/LX columns:', list(zip(LCH_IDX, LX_IDX)))

created = updated = skipped = failed = 0
Garment = env['vbs.garment']
Move = env['vbs.garment.move']

for row_idx, row in enumerate(
    ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
):
    try:
        customer = row[IDX['customer']] if IDX['customer'] >= 0 else None
        if not customer or not str(customer).strip():
            continue  # skip empty rows

        partner = find_or_create_partner(env, customer)
        order = find_order_for_partner(env, partner)

        g_type = map_value(
            row[IDX['type']] if IDX['type'] >= 0 else None,
            GARMENT_TYPE_MAP, default='khac', label='garment_type',
        )
        detail = (row[IDX['detail']] if IDX['detail'] >= 0 else '') or ''
        state = map_value(
            row[IDX['state']] if IDX['state'] >= 0 else None,
            GARMENT_STATE_MAP, default='nhap', label='state',
        )
        location = map_value(
            row[IDX['location']] if IDX['location'] >= 0 else None,
            GARMENT_LOCATION_MAP, default='cua_hang', label='location',
        )
        date_return = to_date(row[IDX['date_return']]) if IDX['date_return'] >= 0 else False

        # Idempotency: tìm garment trùng (same partner + same detail + type)
        existing = Garment.search([
            ('order_id.partner_id', '=', partner.id),
            ('garment_type', '=', g_type),
            ('detail', '=', detail),
        ], limit=1)

        vals = {
            'order_id': order.id,
            'garment_type': g_type,
            'detail': detail,
            'state': state,
            'location': location,
            'date_return': date_return,
        }

        if existing:
            existing.write(vals)
            garment = existing
            updated += 1
        else:
            garment = Garment.create(vals)
            created += 1

        # Import LCH/LX moves. LCH = Cửa hàng → Xưởng; LX = Xưởng → Cửa hàng.
        # Nếu đã có sẵn move cho garment này thì skip (tránh duplicate khi re-run).
        if not Move.search_count([('garment_id', '=', garment.id)]):
            for lch_i, lx_i in zip(LCH_IDX, LX_IDX):
                lch_date = to_date(row[lch_i]) if lch_i >= 0 else False
                lx_date = to_date(row[lx_i]) if lx_i >= 0 else False
                if lch_date:
                    Move.create({
                        'garment_id': garment.id,
                        'move_type': 'lch',
                        'move_date': datetime.combine(lch_date, datetime.min.time()),
                    })
                if lx_date:
                    Move.create({
                        'garment_id': garment.id,
                        'move_type': 'lx',
                        'move_date': datetime.combine(lx_date, datetime.min.time()),
                    })
    except Exception as e:
        failed += 1
        print(f'  [FAIL] row {row_idx}: {e}')

wb.close()
print(f'\n=== KẾT QUẢ ===')
print(f'  created: {created}')
print(f'  updated: {updated}')
print(f'  failed:  {failed}')
print(f'  skipped: {skipped}')

if DRY_RUN:
    print('\n>>> DRY_RUN=True — rollback')
    env.cr.rollback()
else:
    env.cr.commit()
    print('\n>>> committed')
