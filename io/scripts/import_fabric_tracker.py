# -*- coding: utf-8 -*-
"""Import sheet "Theo dõi đặt vải" → vbs.fabric.order + line.

Mỗi row Excel = 1 đơn vải riêng (vì không có group theo Mã Sapo).
Nếu muốn gộp nhiều row cùng Mã Sapo thành 1 phiếu, chỉnh GROUP_BY_SAPO = True.

Usage:
    docker exec -i custom_addons-odoo-1 odoo shell \
        -c /etc/odoo/odoo.conf -d VBS_ERP --no-http \
        < io/scripts/import_fabric_tracker.py
"""
import openpyxl
exec(open('/mnt/io/scripts/_lib.py', encoding='utf-8').read())

INPUT_FILE = '/mnt/io/input/nhat_ky.xlsx'
SHEET_NAME = 'Theo dõi đặt vải'
HEADER_ROW = 2
DATA_START_ROW = 3
GROUP_BY_SAPO = True   # True = gộp các row cùng Mã Sapo vào 1 phiếu
DRY_RUN = False

wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws = wb[SHEET_NAME]

header = [str(c.value).strip() if c.value else '' for c in ws[HEADER_ROW]]
print('HEADER:', header[:15])


def col(name):
    for i, h in enumerate(header):
        if h.lower() == name.lower():
            return i
    return -1


IDX = {
    'sapo': col('Mã Sapo'),
    'customer': col('Tên khách hàng'),
    'brand': col('Hãng vải'),
    'type': col('Chủng loại'),
    'qty': col('Số lượng'),
    'date_order': col('Ngày lên đơn'),
    'date_arrived': col('Ngày vải về'),
    'status': col('Trạng thái'),
    'note': col('Ghi chú'),
}
print('Column map:', IDX)

# Group rows theo Mã Sapo nếu bật flag
groups = {}  # key = sapo (hoặc row_idx nếu không có sapo); value = list of rows
for row_idx, row in enumerate(
    ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
):
    customer = row[IDX['customer']] if IDX['customer'] >= 0 else None
    if not customer or not str(customer).strip():
        continue
    sapo = row[IDX['sapo']] if IDX['sapo'] >= 0 else None
    key = str(sapo).strip() if (GROUP_BY_SAPO and sapo) else f'row_{row_idx}'
    groups.setdefault(key, []).append(row)

print(f'Gom {len(groups)} nhóm phiếu vải từ sheet')

created_orders = created_lines = skipped = failed = 0
Order = env['vbs.fabric.order']
Line = env['vbs.fabric.order.line']

for key, rows in groups.items():
    try:
        first = rows[0]
        customer = first[IDX['customer']]
        partner = find_or_create_partner(env, customer)

        date_order = to_date(first[IDX['date_order']]) if IDX['date_order'] >= 0 else False
        date_arrived = to_date(first[IDX['date_arrived']]) if IDX['date_arrived'] >= 0 else False
        state = map_value(
            first[IDX['status']] if IDX['status'] >= 0 else None,
            FABRIC_STATE_MAP, default='cho_ve', label='fabric state',
        )

        # Idempotency: tìm order đã có với cùng sapo + partner ở dòng đầu
        sapo = (str(first[IDX['sapo']]).strip() if first[IDX['sapo']] else '')
        existing_order = False
        if sapo:
            existing_order = Order.search([
                ('line_ids.sapo_code', '=', sapo),
                ('line_ids.partner_id', '=', partner.id),
            ], limit=1)

        if existing_order:
            skipped += 1
            continue

        order = Order.create({
            'date_order': date_order or False,
            'date_arrived': date_arrived or False,
            'state': state,
        })
        created_orders += 1

        for row in rows:
            sapo_i = str(row[IDX['sapo']]).strip() if (IDX['sapo'] >= 0 and row[IDX['sapo']]) else ''
            brand = (row[IDX['brand']] if IDX['brand'] >= 0 else '') or ''
            type_str = (row[IDX['type']] if IDX['type'] >= 0 else '') or ''
            qty = row[IDX['qty']] if IDX['qty'] >= 0 else 0
            note = (row[IDX['note']] if IDX['note'] >= 0 else '') or ''
            Line.create({
                'order_id': order.id,
                'partner_id': partner.id,
                'sapo_code': sapo_i,
                'fabric_brand': str(brand).strip()[:64],
                'garment_ref': '',
                'quantity': float(qty) if qty else 0.0,
                'note': str(note).strip()[:255],
            })
            created_lines += 1
    except Exception as e:
        failed += 1
        print(f'  [FAIL] group {key}: {e}')

wb.close()
print(f'\n=== KẾT QUẢ ===')
print(f'  orders created: {created_orders}')
print(f'  lines created:  {created_lines}')
print(f'  skipped (dup):  {skipped}')
print(f'  failed:         {failed}')

if DRY_RUN:
    env.cr.rollback()
    print('\n>>> DRY_RUN=True — rollback')
else:
    env.cr.commit()
    print('\n>>> committed')
